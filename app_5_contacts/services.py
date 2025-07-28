from types import NoneType
from tempfile import NamedTemporaryFile
import openpyxl


class Companies:
    def __init__(self, _list):
        self._dict = {c['ID']: c['TITLE'] for c in _list}

    def __getitem__(self, key):
        return self._dict[key]

    def __repr__(self):
        return str(self._dict)

    def by_name(self, search_name):
        for uid, name in self._dict.items():
            if name == search_name:
                return uid
        return None

    def add(self, uid, data):
        self._dict[str(uid)] = data


contacts_format = {
    list: lambda c: [k['VALUE'] for k in c],
    NoneType: lambda c: '',
    str: lambda c: c,
}
class Contacts:
    def __init__(self, _list):
        self._dict = {c['ID']: {t: contacts_format[type(c[t])](c[t])
                                for t in c if t != 'ID'}
                      for c in _list}
        for cid in self._dict:
            self._dict[cid]['COMPANY'] = []
            if 'PHONE' not in self._dict[cid]:
                self._dict[cid]['PHONE'] = []
            if 'EMAIL' not in self._dict[cid]:
                self._dict[cid]['EMAIL'] = []

    def __getitem__(self, key):
        return self._dict[key]

    def __repr__(self):
        return str(self._dict)

    def all(self):
        return tuple(self._dict.keys())

    def by_name(self, search_name, search_last_name):
        for uid, info in self._dict.items():
            if search_name == info['NAME'] and search_last_name == info['LAST_NAME']:
                return uid
        return None

    def as_csv(self):
        file = NamedTemporaryFile('w+b')
        file.write('Имя,Фамилия,Телефон,Почта,Компания\n'.encode('utf-8'))
        for contact in self._dict.values():
            file.write((','.join((contact["NAME"],
                                  contact["LAST_NAME"],
                                  ";".join(contact["PHONE"]),
                                  ";".join(contact["EMAIL"]),
                                  ";".join(contact["COMPANY"]))) + '\n').encode('utf-8'))
        file.seek(0)
        return file

    def as_xlsx(self):
        file = NamedTemporaryFile('w+b')
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(('Имя', 'Фамилия', 'Телефон', 'Почта', 'Компания'))
        for contact in self._dict.values():
            sheet.append((contact["NAME"],
                          contact["LAST_NAME"],
                          ";".join(contact["PHONE"]),
                          ";".join(contact["EMAIL"]),
                          ";".join(contact["COMPANY"])))

        workbook.save(file.name)
        return file

    def add(self, uid, data):
        self._dict[str(uid)] = data


def collect_companies(contacts, companies, search_results, contacts_order):
    for cid, result in zip(contacts_order, search_results):
        for company in result:
            contacts[cid]['COMPANY'].append(companies[str(company['COMPANY_ID'])])


def check_companies_and_contacts(but):
    companies = Companies(but.call_list_method('crm.company.list', fields={'select': ['ID', 'TITLE']}))
    contacts = Contacts(
        but.call_list_method('crm.contact.list', fields={'select': ['ID', 'NAME', 'LAST_NAME', 'PHONE', 'EMAIL']}))

    contacts_order = contacts.all()
    contact2company = but.batch_api_call([('crm.contact.company.items.get', {'id': uid}) for uid in contacts_order])
    contact2company = [i['result'] for i in contact2company.values()]
    collect_companies(contacts, companies, contact2company, contacts_order)
    return companies, contacts


class Parse:
    def __init__(self, file_stream):
        file_format = file_stream.name.split('.')[-1]
        self.user_data = []
        if file_format == 'csv':
            self.parse_as_csv(file_stream)
        elif file_format == 'xlsx':
            self.parse_as_xlsx(file_stream)
        else:
            return

    def parse_as_csv(self, stream):
        row = stream.readline()
        while True:
            row = stream.readline().decode('utf-8').strip()
            if not row:
                break
            name, last_name, phone, email, company = row.split(',')
            self.user_data.append({'NAME': name,
                                   'LAST_NAME': last_name,
                                   'PHONE': phone.split(';'),
                                   'EMAIL': email.split(';'),
                                   'COMPANY': company.split(';') if company else []})

    def parse_as_xlsx(self, stream):
        workbook = openpyxl.load_workbook(stream)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2):
            name, last_name, phone, email, company = [cell.value if cell.value else '' for cell in row]
            self.user_data.append({'NAME': name,
                                   'LAST_NAME': last_name,
                                   'PHONE': phone.split(';'),
                                   'EMAIL': email.split(';'),
                                   'COMPANY': company.split(';') if company else []})

    def differ(self, contacts, companies):
        new_contacts = []
        contacts_updates = {}
        new_companies = []
        for contact in self.user_data:
            exists_id = contacts.by_name(contact['NAME'], contact['LAST_NAME'])
            if exists_id:
                contacts_updates[exists_id] = {}
                for field in ('PHONE', 'EMAIL', 'COMPANY'):
                    for item in contact[field]:
                        if item and item not in contacts[exists_id][field]:
                            if not field in contacts_updates[exists_id]:
                                contacts_updates[exists_id][field] = []
                            contacts_updates[exists_id][field].append(item)
                if not contacts_updates[exists_id]:
                    del contacts_updates[exists_id]
            else:
                new_contacts.append(contact)

            for company in contact['COMPANY']:
                exists_id = companies.by_name(company)
                if not exists_id:
                    new_companies.append(company)
        return new_contacts, new_companies, contacts_updates